r"""
سیستم تطبیق فیش‌های بانکی
===========================================
FastAPI backend + pdfplumber + pandas

نحوه اجرا:
  cd d:\Checkhesab\receipt-checker
  python app.py

سپس مرورگر را باز کنید:
  http://localhost:8765
"""

from __future__ import annotations
import io, re, json, logging, sys
from pathlib import Path
from typing import Optional
from difflib import SequenceMatcher

def get_base_path() -> Path:
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    return Path(__file__).parent

BASE_DIR = get_base_path()

import uvicorn
import pdfplumber
import pandas as pd
import requests
import urllib3
import ssl
from requests.adapters import HTTPAdapter
from urllib3.util.ssl_ import create_urllib3_context
from fastapi import FastAPI, UploadFile, File, Form, Body, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, RedirectResponse, Response
from starlette.middleware.sessions import SessionMiddleware
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("receipt_checker")

app = FastAPI(title="Receipt Checker")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.add_middleware(SessionMiddleware, secret_key="super_secret_checkhesab_key_123")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

ADMIN_USER = "admin"
ADMIN_PASS = "admin"

def check_auth(request: Request):
    if not request.session.get("authenticated"):
        raise HTTPException(status_code=401, detail="Unauthorized")

# ──────────────────────────────────────────────────────────────────────────────
# Server-side Excel Session Storage (per user session)
# ──────────────────────────────────────────────────────────────────────────────
EXCEL_SESSIONS: dict = {}

def _session_key(request: Request) -> str:
    """Return a stable key for this user's session (uses session cookie value)."""
    return request.session.get("session_id", "anonymous")

def _ensure_session_id(request: Request):
    """Create a session_id if not present."""
    if "session_id" not in request.session:
        import uuid
        request.session["session_id"] = str(uuid.uuid4())

# ──────────────────────────────────────────────────────────────────────────────
# Number utilities (Persian/Arabic/English digits + comma removal)
# ──────────────────────────────────────────────────────────────────────────────
FA_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

# ──────────────────────────────────────────────────────────────────────────────
# Date utilities — Shamsi / Gregorian same-day or next-day check
# ──────────────────────────────────────────────────────────────────────────────

def _parse_date_parts(date_str: str):
    """Parse a date string like '1404/01/10' or '1404-01-10' or '14040110'
    into (year, month, day) integers. Returns None if unparseable."""
    if not date_str:
        return None
    s = str(date_str).translate(FA_DIGITS).strip()
    # Try separator-based formats
    for sep in ('/', '-', '.'):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    return (int(parts[0]), int(parts[1]), int(parts[2]))
                except ValueError:
                    pass
    # Try compact 8-digit YMD
    digits = re.sub(r'\D', '', s)
    if len(digits) == 8:
        try:
            return (int(digits[:4]), int(digits[4:6]), int(digits[6:]))
        except ValueError:
            pass
    return None

def _date_ok(receipt_date_str: str, bank_date_str: str) -> bool:
    """Return True if bank_date == receipt_date OR bank_date == receipt_date + 1 day.
    Returns True when either date is missing/unparseable (non-blocking)."""
    rd = _parse_date_parts(receipt_date_str)
    bd = _parse_date_parts(bank_date_str)
    if not rd or not bd:
        return True  # not enough info — allow the match
    # Must be same year/month (any reasonable month), just check day difference
    # We'll use a simple arithmetic: convert to a scalar YYYYMMDD int
    def scalar(t): return t[0] * 10000 + t[1] * 100 + t[2]
    diff = scalar(bd) - scalar(rd)
    # diff == 0: same day; diff == 1: next day  (ignores month-boundary edge cases — close enough for bank matching)
    # Handle month boundary: if receipt is last day of month
    # Simple approach: parse as Python date, but we avoid importing datetime for Shamsi
    # Instead: allow diff == 0 or diff == 1 at same YYYYMM, OR allow day==1 of next month when receipt day >= 28
    if rd[0] == bd[0] and rd[1] == bd[1]:
        return bd[2] - rd[2] in (0, 1)
    # Month boundary: receipt is last day of month, bank is 1st of next month
    if rd[0] == bd[0] and bd[1] == rd[1] + 1 and bd[2] == 1:
        return True
    # Year boundary: receipt=12/29 or 12/30, bank=1/1 of next year
    if bd[0] == rd[0] + 1 and bd[1] == 1 and bd[2] == 1 and rd[1] == 12:
        return True
    return False

# Codes that identify the account holder themselves (not tracking codes)
# These are extracted from the PDF header and should be excluded from per-row matching
ACCOUNT_HOLDER_CODES: set[str] = set()

def to_num(s: str) -> Optional[float]:
    """Convert a possibly-Persian number string to float. Returns None on failure."""
    if not s:
        return None
    s_str = str(s).translate(FA_DIGITS)
    # Strip absolutely everything except digits and decimal point
    clean_s = re.sub(r'[^\d.]', '', s_str)
    if not clean_s:
        return None
    try:
        return float(clean_s)
    except ValueError:
        return None

def clean_str(s) -> str:
    if s is None:
        return ""
    return str(s).strip()

def nrm(s: str) -> str:
    """Normalize Persian/Arabic string for comparison."""
    if not s:
        return ""
    s = s.replace("\u200c", " ").replace("\u200d", " ")  # ZWNJ
    s = s.replace("ك", "ک").replace("ي", "ی")             # Arabic variants
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# ──────────────────────────────────────────────────────────────────────────────
# PDF Parser — pdfplumber table extraction
# ──────────────────────────────────────────────────────────────────────────────

# Column header patterns (right→left reading order in visual PDF)
# The PDF has these columns (left side in x-coordinate space → right side visually):
#   مانده ریالی | بستانکار مالی | بدهکار مالی | مانده طلایی | بستانکار طلا | بدهکار طلا
#   | شرح | شماره سند | تاریخ | حساب
COLUMN_PATTERNS = {
    "حساب":          re.compile(r"حساب"),
    "تاریخ":         re.compile(r"تاریخ"),
    "شماره_سند":     re.compile(r"شماره\s*سند"),
    "شرح":           re.compile(r"^شرح$"),
    "بدهکار_طلا":    re.compile(r"بدهکار\s*طلا"),
    "بستانکار_طلا":  re.compile(r"بستانکار\s*طلا"),
    "مانده_طلا":     re.compile(r"مانده\s*طلا"),
    "بدهکار_مالی":   re.compile(r"بدهکار\s*مالی"),
    "بستانکار_مالی": re.compile(r"بستانکار\s*مالی"),
    "مانده_ریالی":   re.compile(r"مانده\s*ریالی"),
}

def identify_col(header_text: str) -> str:
    """Map a header cell text to a column key."""
    t = clean_str(header_text)
    for key, pat in COLUMN_PATTERNS.items():
        if pat.search(t):
            return key
    return ""

import fitz  # PyMuPDF

def fix_rtl(text: str) -> str:
    """Reverse extracted RTL text and restore LTR layout for numbers and English words."""
    if not text: return ""
    rev = text[::-1]
    return re.sub(r'[A-Za-z0-9/\.,\-_\[\]\(\)]+', lambda m: m.group(0)[::-1], rev)

def parse_pdf(pdf_bytes: bytes) -> list[dict]:
    """Extract rows from the Tahesab account statement PDF using visual word assembly."""
    global ACCOUNT_HOLDER_CODES
    rows_out = []
    ACCOUNT_HOLDER_CODES = set()

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    logger.info(f"PDF pages: {len(doc)}")
    
    for page_num, page in enumerate(doc, 1):
        words = page.get_text("words")
        
        # Extract account holder code from page 1 header (format: Name)CODE( AccountNumber in RTL)
        if page_num == 1:
            for w in words:
                text = w[4]
                # PyMuPDF often reverses RTL text: so (8181) appears as )8181(
                m = re.match(r"^\)(\d{4,5})\($", text)  # e.g. ")8181("
                if not m:
                    m = re.match(r"^\((\d{4,5})\)$", text)  # also check normal form
                if m:
                    ACCOUNT_HOLDER_CODES.add(m.group(1))
                    logger.info(f"Account holder code detected: {m.group(1)}")
        
        rows_map = {}
        
        # 1. Group words by visual Y-row
        for w in words:
            x0, y0, x1, y1, text, b, l, w_num = w
            y_mid = (y0 + y1) / 2
            y_key = round(y_mid / 5) * 5
            if y_key not in rows_map:
                rows_map[y_key] = []
            rows_map[y_key].append((x0, text))

        # 2. Assemble lines by sorting right-to-left (X descending)
        sorted_y = sorted(rows_map.keys())
        for y in sorted_y:
            # sort words by X right-to-left
            row_words = sorted(rows_map[y], key=lambda item: item[0], reverse=True)
            line = " ".join(item[1] for item in row_words).strip()
            
            if not line:
                continue

            # --- Apply Regex Logic to Assembled Line ---
            line_norm = line.translate(FA_DIGITS)
            line_no_spaces = re.sub(r"\s+", "", line_norm)
            
            # Identify valid row: must have a date-like string or a long dash-surrounded document number
            # the garbled dates look like 1/11//1/1 sometimes, but amounts are the real indicator
            
            # Amounts: match sequences of digits with standard commas/spaces
            # Note: PyMuPDF maps the zero glyph to '2' in this PDF font
            # So we look at the raw extracted digits and fix trailing 222... groups
            amounts = []
            for m in re.finditer(r"\b(\d{1,3}(?:[,،./]\d{3})+|\d{5,})\b", line_norm):
                raw = m.group(1).replace(",", "").replace("،", "").replace(".", "").replace("/", "")
                # Fix trailing run of 2s that is actually zeros (PDF font encoding bug)
                # Only replace '22...' at the trailing end if they form groups of 2+
                # Be conservative: only replace trailing 2s if the number has > 6 digits
                fixed = raw
                if len(raw) > 6 and raw.endswith("22"):
                    # Count trailing 2s
                    trailer = len(raw) - len(raw.rstrip("2"))
                    if trailer >= 2:
                        fixed = raw[:-trailer] + "0" * trailer
                amt = to_num(fixed)
                if amt and amt > 1000 and len(fixed) <= 9:  # exclude 10+ digit account numbers
                    amounts.append(amt)
            
            amounts.sort(reverse=True)
            
            # If no large amounts, it's not a transaction data row
            if not amounts:
                continue
                
            # Usually credit is the 2nd largest number, or 1st if there's only one
            credit = amounts[1] if len(amounts) >= 2 else amounts[0]
            
            codes, sender = parse_desc(line)
            
            desc_fixed = line[:200].strip()

            doc_m = re.search(r"\b(\d{6,8})\b", line_norm)
            doc_num = doc_m.group(1) if doc_m else ""
            
            # Since dates can be weird, we just match a naive pattern or leave empty
            date_m = re.search(r"1[34]\d\d[\D]\d{1,2}[\D]\d{1,2}", line_no_spaces)
            date_str = date_m.group(0).replace('-', '/').replace('_', '/') if date_m else ""

            if credit > 0:
                rows_out.append({
                    "page":        page_num,
                    "date":        date_str,
                    "doc_num":     doc_num,
                    "desc":        desc_fixed,
                    "credit":      credit,
                    "debit":       0,
                    "credit_raw":  str(credit),
                    "debit_raw":   "0",
                    "codes":       codes,
                    "sender":      sender,
                    "doc_type":    "بستانکار",
                    "amount":      credit,
                })

    doc.close()
    logger.info(f"Total PDF rows extracted: {len(rows_out)}")
    
    # Auto-detect and filter codes that appear in >30% of all rows
    # These are "owner codes" (like account branch codes) that appear everywhere
    if rows_out:
        code_freq: dict[str, int] = {}
        for r in rows_out:
            for c in r.get("codes", []):
                code_freq[c] = code_freq.get(c, 0) + 1
        threshold = max(5, len(rows_out) * 0.1)
        auto_owner_codes = {c for c, cnt in code_freq.items() if cnt > threshold}
        if auto_owner_codes:
            logger.info(f"Auto-detected owner codes (filtered from tracking): {auto_owner_codes}")
            ACCOUNT_HOLDER_CODES.update(auto_owner_codes)
            for r in rows_out:
                r["codes"] = [c for c in r["codes"] if c not in ACCOUNT_HOLDER_CODES]

    return rows_out

def parse_html(html_bytes: bytes) -> list[dict]:
    """Parse HTML statement into the standard dict format."""
    
    # --- Smart charset detection ---
    # Many Iranian accounting tools (e.g. Tahesab) save HTML with windows-1256.
    # We must detect the encoding from the <meta charset> tag BEFORE decoding,
    # because windows-1256 bytes are valid latin-1 and won't raise UnicodeDecodeError
    # when decoded as UTF-8 — they just silently produce ??? for all Persian text.
    import re as _re
    # Step 1: try to sniff charset from raw bytes using a quick latin-1 decode
    _sniff = html_bytes[:2000].decode('latin-1', errors='replace')
    _charset_match = _re.search(r'charset\s*=\s*["\']?\s*([\w-]+)', _sniff, _re.IGNORECASE)
    _detected_enc = _charset_match.group(1).strip().lower() if _charset_match else 'utf-8'
    # Normalize common aliases
    if _detected_enc in ('windows-1256', 'cp1256', '1256', 'arabic'):
        _detected_enc = 'cp1256'
    elif _detected_enc in ('windows-1252', 'cp1252', 'iso-8859-1', 'latin-1'):
        _detected_enc = 'cp1252'
    # Step 2: decode with the detected encoding, fallback to utf-8 then cp1256
    try:
        html_text = html_bytes.decode(_detected_enc)
    except (UnicodeDecodeError, LookupError):
        try:
            html_text = html_bytes.decode('utf-8')
        except UnicodeDecodeError:
            html_text = html_bytes.decode('cp1256', errors='replace')
        
    soup = BeautifulSoup(html_text, "html.parser")
    rows_out = []
    seen_tx = set()
    
    
    # Try to find exactly which columns contain the desc/credit/debit
    # by looking at table headers
    desc_idx = credit_idx = debit_idx = date_idx = -1
    
    tables = soup.find_all("table")
    for table in tables:
        rows_all = table.find_all("tr")
        if not rows_all:
            continue
        
        # Detect column indices from the FIRST header row (not flattened)
        header_cells = rows_all[0].find_all(["th", "td"])
        desc_indices = []
        credit_idx = -1
        debit_idx = date_idx = doc_num_idx = account_name_idx = tracking_idx = -1
        for ci, cell in enumerate(header_cells):
            # Normalize Arabic Yeh/Kaf to Persian so our "in" checks work reliably
            text = cell.get_text(strip=True).replace('ي', 'ی').replace('ك', 'ک')
            if "شرح" in text or "نوع" in text:
                desc_indices.append(ci)
            elif "بستانکار" in text and "مالی" in text:
                credit_idx = ci
            elif "بدهکار" in text and "مالی" in text:
                debit_idx = ci
            elif "تاریخ" in text:
                if date_idx == -1:
                    date_idx = ci          # Take first تاریخ column as default
                if "عملیات" in text:
                    date_idx = ci          # Override with تاریخ عملیات if found
            elif "نام حساب" in text or "نامحساب" in text or "صاحب حساب" in text:
                account_name_idx = ci
            elif ("شماره" in text and "سند" in text) or "سماره" in text:
                doc_num_idx = ci
            elif "رهگیری" in text or "رسید" in text:
                tracking_idx = ci
        
        # Fallbacks if بستانکار مالی / بدهکار مالی not matched
        if credit_idx == -1 or debit_idx == -1:
            for ci, cell in enumerate(header_cells):
                text = cell.get_text(strip=True).replace('ي', 'ی').replace('ك', 'ک')
                if "بستانکار" in text and credit_idx == -1:
                    credit_idx = ci
                elif "بدهکار" in text and debit_idx == -1:
                    debit_idx = ci

        # If we found at least credit in this table, process its rows
        if credit_idx != -1:
            for tr in rows_all[1:]:  # Skip header row
                tds = tr.find_all("td")
                if not tds or len(tds) <= credit_idx:
                    continue

                # Build desc ONLY from the شرح column if it has a dedicated column,
                # otherwise join all cells EXCEPT the structured columns (credit, debit, date, account_name, doc_num).
                # This prevents account holder name / amounts from polluting the sender extraction.
                excluded_cols = {i for i in [credit_idx, debit_idx, date_idx, account_name_idx, doc_num_idx] if i != -1}
                
                if desc_indices:
                    desc_parts = []
                    for di in desc_indices:
                        if di < len(tds):
                            part = tds[di].get_text(separator=" ", strip=True).replace('ي', 'ی').replace('ك', 'ک')
                            if part and part != "-":
                                desc_parts.append(part)
                    desc_text = " - ".join(desc_parts) if desc_parts else ""
                else:
                    desc_parts_filtered = [tds[ci].get_text(strip=True) for ci in range(len(tds))
                                           if ci not in excluded_cols and tds[ci].get_text(strip=True)]
                    desc_text = " | ".join(desc_parts_filtered)

                # Append tracking code column explicitly so Regex can find the 4-digit code
                if tracking_idx != -1 and len(tds) > tracking_idx and tracking_idx not in desc_indices:
                    tk_text = tds[tracking_idx].get_text(strip=True)
                    if tk_text and tk_text != "-":
                        desc_text += f" - {tk_text}"
                
                desc = desc_text.replace('ي', 'ی').replace('ك', 'ک')

                first_col = tds[0].get_text(strip=True) if tds else ""
                if "ردیف" in first_col or not desc:
                    continue

                credit_raw = tds[credit_idx].get_text(strip=True)
                debit_raw  = tds[debit_idx].get_text(strip=True)  if debit_idx != -1 and len(tds) > debit_idx  else "0"
                date_str   = tds[date_idx].get_text(strip=True)   if date_idx  != -1 and len(tds) > date_idx   else ""
                # Normalize date: remove leading day-letter prefix (like 'د ', 'ش ', etc.)
                date_str = re.sub(r'^[آابپتثجچحخدذرزژسشصضطظعغفقکگلمنوهی]\s+', '', date_str).strip()
                
                # Extract document number from dedicated column if available
                doc_num_str = tds[doc_num_idx].get_text(strip=True) if doc_num_idx != -1 and len(tds) > doc_num_idx else ""
                
                credit = to_num(credit_raw) or 0
                debit = to_num(debit_raw) or 0
                
                if credit == 0 and debit == 0:
                    continue # Not a transaction row
                    
                # Improved duplicate detection: hash the ENTIRE row text, not just the first 100 chars, 
                # because 17-column tables start with the same data for many rows.
                tx_key = (date_str, desc, credit, debit)
                if tx_key in seen_tx:
                    continue
                seen_tx.add(tx_key)
                    
                codes, parsed_sender = parse_desc(desc)
                
                account_name = ""
                if account_name_idx != -1 and len(tds) > account_name_idx:
                    account_name = tds[account_name_idx].get_text(strip=True).replace('ي', 'ی').replace('ك', 'ک').strip()
                
                rows_out.append({
                    "page":          1,
                    "date":          date_str,
                    "doc_num":       doc_num_str,
                    "desc":          desc[:200],
                    "credit":        credit,
                    "debit":         debit,
                    "credit_raw":    credit_raw,
                    "debit_raw":     debit_raw,
                    "codes":         codes,
                    "sender":        parsed_sender,   # person who sent the money (from description)
                    "customer_name": account_name,   # account holder name (نام حساب)
                    "doc_type":      "بستانکار" if credit > 0 else "بدهکار",
                    "amount":        credit if credit > 0 else debit,
                })
            # Do not break here! Process all tables because HTML might be paginated
            # with multiple transaction tables.
            pass
                
    logger.info(f"Total HTML rows extracted: {len(rows_out)}")
    
    # Auto-detect owner codes as we do for PDF
    if rows_out:
        code_freq: dict[str, int] = {}
        for r in rows_out:
            for c in r.get("codes", []):
                code_freq[c] = code_freq.get(c, 0) + 1
        threshold = max(5, len(rows_out) * 0.1)
        auto_owner_codes = {c for c, cnt in code_freq.items() if cnt > threshold}
        if auto_owner_codes:
            global ACCOUNT_HOLDER_CODES
            ACCOUNT_HOLDER_CODES.update(auto_owner_codes)
            for r in rows_out:
                r["codes"] = [c for c in r["codes"] if c not in ACCOUNT_HOLDER_CODES]

    return rows_out

def parse_desc(desc: str) -> tuple[list[str], str]:
    """
    Extract 4-digit tracking codes and sender name from شرح column.

    Examples:
      "واریز نقد به بانک (از مشتری) [صادرات منشادی،24454] خدمتی/2452"
      "واریز نقد به بانک (از مشتری) [100617] واحدشه/1264"
      "واریز نقد به بانک (از مشتری) [21102] برهام/1102"
      "[97830/3882] دادور/2951"
    """
    if not desc:
        return [], ""

    # Normalize Persian/Arabic digits to Latin
    desc_n = desc.translate(FA_DIGITS)

    codes: set[str] = set()

    # Pattern 1: فارسی/XXXX  — tracking code after sender name + slash
    # Pattern 1a: فارسی/XXXX — sender name + slash + tracking code
    for m in re.finditer(r"[\u0600-\u06FF]+\s*/\s*(\d{4,15})\b", desc_n):
        codes.add(m.group(1))

    # Pattern 1b: XXXX/فارسی  — tracking code before sender name + slash
    for m in re.finditer(r"\b(\d{4,15})\s*/\s*[\u0600-\u06FF]+", desc_n):
        codes.add(m.group(1))

    # Pattern 2: NNNN[اسم فارسی] — 4-digit code BEFORE a bracket (e.g. 8842[غفاری])
    for m in re.finditer(r"(?<!\d)(\d{4,15})(?!\d)\s*\[", desc_n):
        codes.add(m.group(1))

    # Pattern 3: inside brackets [بانک،NNNNN]
    for m in re.finditer(r"\[([^\]]+)\]", desc_n):
        inner = m.group(1)
        nums = re.findall(r"\d+", inner)
        for n in nums:
            if len(n) >= 4:
                codes.add(n)   # Keep full number

    # Pattern 4: code AFTER closing bracket ] (PyMuPDF reverses RTL brackets)
    # Format: "[ بانک منشادی ، ]3030," → code is 3030 (right after ])
    # Only match if the bracket content is long (bank name), NOT short person names
    for m in re.finditer(r"\[([^\]]{8,})\]\s*(\d{4,15})\b", desc_n):
        codes.add(m.group(2))

    # Pattern 5: Isolated numbers surrounded by | spaces |
    for m in re.finditer(r"\|\s*(\d{4,15})\s*(?=\|)", desc_n):
        codes.add(m.group(1))

    # Pattern 6: Isolated numbers at the end of the line (especially added from tracking col via ' - ')
    for m in re.finditer(r"(?:\||-)\s*(\d{4,15})\s*$", desc_n):
        codes.add(m.group(1))

    # Pattern 7: Multiple 4-15 digit numbers separated by slashes/dashes (like 2356/8765/13291)
    for m in re.finditer(r"(?<!\d)(\d{4,15})(?:\s*[/\\-]\s*(\d{4,15}))+(?!\d)", desc_n):
        match_str = m.group(0)
        parts = re.findall(r"\d+", match_str)
        # Only skip if this looks like a full Shamsi date (4-digit year / 1-2 digit month / 1-2 digit day)
        # A real date has year >= 1380, month 1-12, day 1-31
        if len(parts) == 3 and int(parts[0]) >= 1380 and 1 <= int(parts[1]) <= 12 and 1 <= int(parts[2]) <= 31:
            continue  # This is a date like 1404/01/10 — skip it
        # Otherwise extract all parts as codes
        for n in parts:
            if len(n) >= 4:
                codes.add(n)

    # Pattern 8: Broad fallback — any remaining isolated 4-15 digit number not yet extracted
    # This catches codes like '4110' that appear in text like 'پایا - 4110 - نام'
    # without any brackets/slashes. Applied LAST to avoid duplicate work.
    already_found = set(codes)
    for m in re.finditer(r'(?<![\d])(\d{4,15})(?![\d])', desc_n):
        n = m.group(1)
        if n in already_found:
            continue
        nint = int(n)
        # Skip Shamsi year range (1380–1420)
        if 1380 <= nint <= 1420:
            continue
        # Skip Iranian mobile numbers (11 digits starting with 09)
        if len(n) == 11 and n.startswith('09'):
            continue
        # Skip large numbers that are clearly amounts (>= 12 digits or round millions)
        # Wait, if tracking code is 12 digits, it's valid. But amount > 999 million is also 10 digits.
        # Bank tracking columns (Pattern 6 explicit checks) have already caught explicit tracking codes.
        # Fallback shouldn't loosely capture 9+ digits unless we are sure it's tracking. 
        # But to be safe, we allow up to 15 digits here if not mobile or date.
        codes.add(n)

    sender = ""
    # Simplify regex to prevent catastrophic backtracking (ReDoS)
    # Match 1+ Persian words before / 4-digit
    m = re.search(r"([\u0600-\u06FF\s]+)\s*/\s*\d{4}", desc)
    if m:
        sender = m.group(1).strip()
    else:
        m2 = re.search(r"\d{4}\s*/\s*([\u0600-\u06FF\s]+)", desc)
        if m2:
            sender = m2.group(1).strip()
        else:
            # Fallback: any Persian name after last bracket
            m = re.search(r"\]\s*([\u0600-\u06FF\s]+)", desc)
            if m:
                sender = m.group(1).strip()
            else:
                # Fallback: حواله به: Name
                m_havale = re.search(r"حواله\s*به:\s*\d*\s*([\u0600-\u06FF\s]+)", desc)
                if m_havale:
                    sender2 = m_havale.group(1).strip()
                    sender2 = re.sub(r'[^\u0600-\u06FF\s]+$', '', sender2).strip()
                    sender = sender2
                    
                    # Also try to extract the actual sender name appearing BEFORE (حواله به
                    m_before = re.search(r"\|\s*([^\|]+?)\s*\(حواله\s*به:", desc)
                    if m_before:
                        sender1 = m_before.group(1).strip()
                        sender1 = re.sub(r'[^\u0600-\u06FF\s]+', '', sender1).strip()
                        if sender1:
                            sender = f"{sender1} {sender2}"

    # Filter out codes that belong to the account holder (not tracking codes)
    codes -= ACCOUNT_HOLDER_CODES

    return sorted(codes), sender

# ──────────────────────────────────────────────────────────────────────────────
# Excel Parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_excel(excel_bytes: bytes, filename: str) -> list[dict]:
    """
    Parse bank Excel statement. Detects header row automatically.
    Returns list of transactions with: ref, last4, amount, date, desc, sender.
    """
    # Detect engine from extension
    ext = Path(filename).suffix.lower()
    engine = "xlrd" if ext == ".xls" else "openpyxl"

    try:
        df_raw = pd.read_excel(io.BytesIO(excel_bytes), engine=engine, header=None,
                               dtype=str, na_filter=False)
    except Exception as e:
        raise HTTPException(400, f"خطا در خواندن Excel: {e}")

    logger.info(f"Excel raw shape: {df_raw.shape}")

    # Find header row
    header_kw = {
        "ref":    re.compile(r"مرجع|پیگیری|reference|شناسه|ارجاع|trace|شماره\s*سند", re.I),
        "credit": re.compile(r"بستانکار|واریز|credit|دریافتی", re.I),
        "debit":  re.compile(r"بدهکار|برداشت|debit|پرداختی", re.I),
        "amount": re.compile(r"مبلغ|amount", re.I),
        "date":   re.compile(r"تاریخ|date", re.I),
        "desc":   re.compile(r"شرح|توضیح|description|بابت|نوع", re.I),
        "sender": re.compile(r"نام|واریزکننده|صاحب|sender", re.I),
    }

    # Bank Melli often uses "واریز" for amount, "شماره سند" for ref, "شرح" for desc
    
    header_idx = -1
    col_map: dict[str, int] = {}

    for ri in range(min(20, len(df_raw))):
        row = df_raw.iloc[ri]
        hits, tmp = 0, {}
        for ci, cell in enumerate(row):
            c = str(cell).strip()
            # Melli specific columns
            if re.search(r"بستانکار|واریز|credit|دریافتی", c, re.I) and "credit" not in tmp:
                tmp["credit"] = ci; hits += 1
            elif re.search(r"بدهکار|برداشت|debit|پرداختی", c, re.I) and "debit" not in tmp:
                tmp["debit"] = ci; hits += 1
            elif re.search(r"مبلغ|amount", c, re.I) and "amount" not in tmp:
                tmp["amount"] = ci; hits += 1
            elif re.search(r"تاریخ|date", c, re.I) and "date" not in tmp:
                tmp["date"] = ci; hits += 1
            elif re.search(r"شرح|توضیح|description|بابت|نوع", c, re.I) and "desc" not in tmp:
                tmp["desc"] = ci; hits += 1
            elif re.search(r"شماره\s*سند|مرجع|پیگیری|reference|شناسه|ارجاع|trace", c, re.I) and "ref" not in tmp:
                tmp["ref"] = ci; hits += 1
            elif re.search(r"اطلاعات\s*اضافی", c, re.I) and "extra_info" not in tmp:
                tmp["extra_info"] = ci
            elif re.search(r"فیش.*حواله|حواله.*فیش", c, re.I) and "fish" not in tmp:
                tmp["fish"] = ci
                
        if hits >= 3:
            header_idx = ri; col_map = tmp
            logger.info(f"Excel header at row {ri}: {col_map}")
            break

    txns = []

    if header_idx >= 0:
        for ri in range(header_idx + 1, len(df_raw)):
            row = df_raw.iloc[ri]
            def g(k): return clean_str(row.iloc[col_map[k]]) if k in col_map else ""
            
            ref  = g("ref")
            date = g("date")
            desc = g("desc")
            extra_info = g("extra_info")
            fish_info = g("fish")
            
            # Combine Saderat extra info columns
            if extra_info and extra_info.lower() != "nan":
                desc = f"{desc} | {extra_info}"
            if fish_info and fish_info.lower() != "nan":
                desc = f"{desc} | {fish_info}"
            
            desc = desc.strip(" | ")
            
            # Extract sender from description for Melli bank statements
            # Typically something like "-0412060171037205-ملي-حسن-زارع  نريماني"
            sndr = ""
            m = re.search(r"[-\s]([^\-\s\d]{3,}(?:[\s\-][^\-\s\d]{3,})*)$", desc)
            if m:
                sndr = m.group(1).replace("-", " ").strip()
            else:
                sndr = desc
            
            amt = 0
            tx_type = "unknown"
            
            # Determine credit/debit
            credit_amt = to_num(g("credit")) or 0 if "credit" in col_map else 0
            debit_amt = to_num(g("debit")) or 0 if "debit" in col_map else 0
            general_amt = to_num(g("amount")) or 0 if "amount" in col_map else 0
            amt_raw = g("amount") or g("credit") or g("debit")
            
            
            if credit_amt > 0:
                amt = credit_amt
                tx_type = "deposit"
            elif debit_amt > 0:
                amt = debit_amt
                tx_type = "withdrawal"
            elif general_amt > 0:
                amt = general_amt
                # Guess based on description if it's a general amount column
                if re.search(r"واریز|بستانکار|دریافت|اعتبار", desc):
                    tx_type = "deposit"
                elif re.search(r"برداشت|بدهکار|پرداخت|خرید", desc):
                    tx_type = "withdrawal"
            
            if not ref and not amt and not date:
                continue
                
            ref_digits = re.sub(r"\D", "", ref)
            desc_digits = re.sub(r"\D", "", str(desc))
            
            # All possible 4+ digit codes from this bank row:
            bank_codes: set[str] = set()
            
            # (1) Ref string (full length if >= 4)
            if len(ref_digits) >= 4:
                bank_codes.add(ref_digits)
            
            # (2) Digits immediately before English letters (e.g. "2337GPPC")
            for m_eng in re.finditer(r"(\d{4,})[A-Za-z]", ref):
                bank_codes.add(m_eng.group(1))
            
            # (3) Codes inside description
            desc_n = str(desc).translate(FA_DIGITS)
            for m_desc in re.finditer(r"\d{10,}", desc_n):
                bank_codes.add(m_desc.group(0))
            for m_desc in re.finditer(r"\b(\d{4,8})\b", desc_n):
                if m_desc.group(1) != str(amt_raw).translate(FA_DIGITS):
                    bank_codes.add(m_desc.group(1))
                    
            # (4) Scan the entire row for hidden long tracking codes (10+ digits)
            # Some bank exports place the tracking code in unmapped columns (e.g. col 18)
            row_joined = " | ".join(str(c) for c in row if pd.notna(c) and str(c).strip())
            row_n = row_joined.translate(FA_DIGITS)
            for m_long in re.finditer(r"\b(\d{10,30})\b", row_n):
                val = m_long.group(1)
                if val != str(amt_raw).translate(FA_DIGITS):
                    bank_codes.add(val)
            
            # (5) Also extract the digit-only portion of IBAN strings (IR + digits)
            # e.g. "IR570130100000000394773883" → "570130100000000394773883"
            for m_iban in re.finditer(r"\bIR(\d{22,26})\b", row_joined, re.IGNORECASE):
                iban_digits = m_iban.group(1)
                bank_codes.add(iban_digits)
                # Also add last 5 and last 4 for short-code matching
                if len(iban_digits) > 5:
                    bank_codes.add(iban_digits[-5:])
                if len(iban_digits) > 4:
                    bank_codes.add(iban_digits[-4:])
            
            # Primary reference number
            m_eng = re.search(r"(\d{4,})[A-Za-z]", ref)
            last4 = m_eng.group(1) if m_eng else ref_digits
            
            # Check for duplicate lock
            raw_joined = " | ".join(str(r) for r in row if str(r).strip())
            is_locked = False
            lock_text = ""
            for cell in row:
                if "تطبیق شده" in str(cell):
                    is_locked = True
                    lock_text = str(cell).strip()
                    break
            
            txns.append({
                "row_num":   ri + 1,
                "ref":       ref,
                "last4":     last4,
                "all_codes": sorted(bank_codes),   # list, not set — JSON serializable
                "amount":    amt,
                "tx_type":   tx_type,
                "date":      date,
                "desc":      desc,
                "sender":    sndr,
                "raw":       raw_joined,
                "is_locked": is_locked,
                "lock_text": lock_text,
            })
    else:
        # Auto-detect: scan every row for big numbers
        logger.warning("No header found — auto-detecting Excel rows")
        for ri in range(len(df_raw)):
            row = df_raw.iloc[ri]
            joined = " ".join(str(c) for c in row)
            joined_n = joined.translate(FA_DIGITS)
            refs    = re.findall(r"\b(\d{10,24})\b", joined_n)
            amounts = [float(m) for m in re.findall(r"\b(\d{6,})\b", joined_n) if float(m) > 10000]
            if not refs and not amounts:
                continue
            ref = sorted(refs, key=len, reverse=True)[0] if refs else ""
            amt = sorted(amounts, reverse=True)[0] if amounts else None
            # find Persian name
            sndr_m = re.search(r"[\u0600-\u06FF]{4,}(?:\s[\u0600-\u06FF]{3,})*", joined)
            sndr = sndr_m.group(0) if sndr_m else ""
            
            is_locked = False
            lock_text = ""
            for cell in row:
                if "تطبیق شده" in str(cell):
                    is_locked = True
                    lock_text = str(cell).strip()
                    break
            
            tx_type = "unknown"
            if re.search(r"واریز|بستانکار|دریافت", joined_n): tx_type = "deposit"
            elif re.search(r"برداشت|بدهکار|پرداخت", joined_n): tx_type = "withdrawal"
            
            txns.append({
                "row_num": ri + 1,
                "ref": ref, "last4": ref[-4:] if len(ref) >= 4 else ref,
                "all_codes": [ref] if ref else [],
                "amount": amt, "tx_type": tx_type, "date": "", "desc": joined[:80],
                "sender": sndr,
                "raw": joined[:120],
                "is_locked": is_locked,
                "lock_text": lock_text,
            })

    logger.info(f"Excel transactions: {len(txns)}")
    return txns

# ──────────────────────────────────────────────────────────────────────────────
# Matching Engine
# ──────────────────────────────────────────────────────────────────────────────

def match_receipts(
    pdf_rows: list[dict],
    bank_txns: list[dict],
    credit_only: bool = True,
    use_tracking: bool = True,
    use_name: bool = True,
    use_amount: bool = True,
    tx_type_filter: str = "all",
    use_date: bool = False,
    already_used_rows: set = None,
) -> list[dict]:
    """Match PDF receipt rows against bank transactions."""

    # Filter PDF rows — always include rows that have tracking codes even if credit_only is on.
    # A debit row with a tracking code is still a verifiable receipt.
    rows = [r for r in pdf_rows if
            (not credit_only or (r.get("credit") and r["credit"] > 0))
            or r.get("codes")]  # Always include rows that have tracking codes

    # ── 1. Strict Rule: Only keep banking-related rows ──
    # Exclude gold/currency trading only if NO tracking code is present.
    # If a row has a tracking code (≥4 digits), it's a verifiable bank transfer
    # regardless of the description keyword — include it.
    GOLD_TYPES  = re.compile(r"خرید\s*طلا|فروش\s*طلا|خرید\s*ارز|فروش\s*ارز|صرافی|سکه", re.IGNORECASE)
    VALID_DOC_TYPES = re.compile(r"واریز|واريز|خروج|برداشت|حواله|انتقال|ساتنا|پایا|نقد|فیش|چک|دریافت", re.IGNORECASE)

    banking_rows = []
    for r in rows:
        desc_str = str(r.get("desc", "")) + " " + str(r.get("no", ""))
        has_banking_kw = VALID_DOC_TYPES.search(desc_str)
        has_gold_kw    = GOLD_TYPES.search(desc_str)
        has_code       = any(len(c) >= 4 for c in r.get("codes", []))

        if has_gold_kw and not has_code:
            continue  # Pure gold/currency trade with no tracking code — skip
        if has_banking_kw or has_code:
            banking_rows.append(r)
            # (rows with neither keyword nor code are quietly skipped)

    logger.info(f"Rows after filtering for valid banking types: {len(banking_rows)}")

    rows = banking_rows

    logger.info(f"PDF/HTML rows for matching: {len(rows)}")

    # Build bank lookup maps
    by_last4: dict[str, list] = {}
    by_amount: dict[str, list] = {}
    by_sender: dict[str, list] = {}
    iban_code_keys: set = set()  # (code, row_num) pairs that came from IBAN strings

    for tx in bank_txns:
        # Filter by transaction type if specified
        if tx_type_filter != "all" and tx.get("tx_type") != "unknown" and tx.get("tx_type") != tx_type_filter:
            continue
            
        # Index ALL possible codes from this bank row
        all_codes = set(tx.get("all_codes", []))
        if tx.get("last4"):
            all_codes.add(tx["last4"])
            
        seen_keys: set = set()
        for code in all_codes:
            if code:
                # Detect if this code was derived from an IBAN (IR...) — already flagged in all_codes
                # We mark it by checking if the original all_codes list contained IBAN-derived ones
                # (parse_excel adds last4/last5 of IBAN digits to all_codes directly)
                for key in [code] + ([code[-5:]] if len(code) > 5 else []) + ([code[-4:]] if len(code) > 4 else []):
                    if key and (key, tx.get("row_num")) not in seen_keys:
                        by_last4.setdefault(key, []).append(tx)
                        seen_keys.add((key, tx.get("row_num")))
        
        # Track which codes came from IBANs (codes that are exactly 4 or 5 chars and appear as
        # suffix of a 22+ digit number which itself starts with a country-code prefix)
        # We stored these in all_codes — we can detect them by checking if any 22+ digit code
        # has these as suffix
        for c in all_codes:
            if len(c) >= 22:
                sfx4 = c[-4:] if len(c) >= 4 else ""
                sfx5 = c[-5:] if len(c) >= 5 else ""
                if sfx4:
                    iban_code_keys.add((sfx4, tx.get("row_num")))
                if sfx5:
                    iban_code_keys.add((sfx5, tx.get("row_num")))
        if tx.get("amount"):
            k = str(int(tx["amount"]))
            by_amount.setdefault(k, []).append(tx)
        for src in [tx.get("sender", ""), tx.get("desc", "")]:
            k = nrm(src)
            if len(k) >= 3:
                by_sender.setdefault(k, []).append(tx)
                
    if bank_txns:
        logger.info(f"Sample bank txns last4s: {[t.get('last4') for t in bank_txns[:10]]}")

    results = []
    
    if rows:
        logger.info(f"Sample PDF row codes: {[r.get('codes') for r in rows[:10]]}")
        
    for r in rows:
        amount  = r.get("amount") or r.get("credit") or r.get("debit") or 0
        codes   = r.get("codes", [])
        sender  = r.get("sender", "")

        matched = None
        status  = "not_found" # "exact", "review", "not_found"
        method  = ""
        # Date is ALWAYS checked (user requirement: must be same day or next day)
        receipt_date = r.get("date", "")

        # ── 1. Match by tracking code + amount (both are REQUIRED) ──
        # The user requirement: 4/5 digit code from HTML must match last 4/5 digits of the
        # Excel tracking code, AND the amount must also match (within 5% tolerance).
        # Sender name match is a bonus that increases confidence.
        if use_tracking:
            for code in codes:
                if len(code) < 4:
                    continue
                
                # Build search codes: try the full code first, then last-5, then last-4
                search_codes = [code]
                if len(code) > 5:
                    search_codes.append(code[-5:])
                if len(code) > 4:
                    search_codes.append(code[-4:])
                    
                for scode in search_codes:
                    cands = by_last4.get(scode, [])
                    if not cands:
                        continue
                    
                    # REQUIRED: amount MUST match EXACTLY
                    if amount:
                        amount_cands = [c for c in cands if c.get("amount") and c["amount"] == amount]
                        # REQUIRED: Date MUST match (same day or next day)
                        amount_cands = [c for c in amount_cands if _date_ok(receipt_date, c.get("date", ""))]
                    else:
                        amount_cands = []
                    
                    if not amount_cands:
                        continue
                    
                    # BONUS: also check sender name for disambiguation / confidence boost
                    sender_key = nrm(sender)
                    name_cands = []
                    if sender_key and len(sender_key) >= 3:
                        name_cands = [c for c in amount_cands 
                                     if sender_key in nrm(c.get("sender", "")) 
                                     or nrm(c.get("sender", "")) in sender_key]
                    
                    # Pick the best candidates (name matches first, then any amount match)
                    working_cands = name_cands if name_cands else amount_cands
                    
                    if len(working_cands) > 1:
                        # Multiple equally valid rows — flag as ambiguous
                        matched = working_cands[0]
                        method = f"کد {scode} + مبلغ (چند مورد مشابه)"
                        status = "multiple"
                        matched["all_matched_rows"] = [c.get("row_num") for c in working_cands if c.get("row_num")]
                    elif len(working_cands) == 1:
                        locked_cand = working_cands[0] if working_cands[0].get("is_locked") else None
                        is_iban_match = (scode, working_cands[0].get("row_num")) in iban_code_keys
                        iban_note = " (با شماره شبا منطبق است)" if is_iban_match else ""
                        
                        if locked_cand:
                            # It's an exact match, but this Excel row was ALREADY locked by a previous receipt
                            matched = locked_cand
                            status = "duplicate"
                            # Tell the user clearly that this specific receipt is a duplicate of something already processed
                            method = f"کد {scode} + مبلغ{iban_note} (تکراری در فایل اکسل)"
                            matched["duplicate_lock_text"] = matched.get("lock_text", "قبلاً تطبیق شده")
                        else:
                            name_bonus = " + نام" if name_cands else ""
                            matched, method, status = working_cands[0], f"کد {scode} + مبلغ{name_bonus}{iban_note}", "exact"
                    
                    if matched:
                        break
                
                if matched:
                    break

        # ── 2. Match by sender name AND amount (Needs Manual Review) ──
        if not matched and use_name and sender and amount:
            skey = nrm(sender)
            
            # First, check exact matches (or fully contained strings)
            cands = by_sender.get(skey, [])
            if cands:
                # Amount MUST match exactly + Mandatory date filter
                amount_cands = [c for c in cands if c.get("amount") and c["amount"] == amount]
                amount_cands = [c for c in amount_cands if _date_ok(receipt_date, c.get("date", ""))]
                
                if amount_cands:
                    locked_cand = next((c for c in amount_cands if c.get("is_locked")), None)
                    if locked_cand:
                        matched = locked_cand
                        status = "duplicate"
                        method = "نام مشابه + مبلغ یکسان (تکراری)"
                        matched["duplicate_lock_text"] = matched.get("lock_text", "قبلاً تطبیق شده")
                    else:
                        matched = amount_cands[0]
                        method, status = f"نام مشابه + مبلغ یکسان", "review"

            # Partial/Fuzzy name match (at least 50% similarity, and amount matches EXACTLY)
            if not matched and len(skey) >= 3:
                # To maximize performance, initially find any Excel txns with exactly matching amount + date
                target_amount_cands = []
                for _, t in enumerate(bank_txns):
                    if t.get("amount") == amount and _date_ok(receipt_date, t.get("date", "")):
                        target_amount_cands.append(t)
                
                # Check fuzzy match against the sender/desc fields of these matching amount transactions
                for c in target_amount_cands:
                    excel_sender = nrm(c.get("sender", ""))
                    excel_desc = nrm(c.get("desc", ""))
                    
                    # Compute ratio
                    ratio1 = SequenceMatcher(None, skey, excel_sender).ratio() if excel_sender else 0
                    ratio2 = SequenceMatcher(None, skey, excel_desc).ratio() if excel_desc else 0
                    
                    if ratio1 >= 0.5 or ratio2 >= 0.5 or (len(excel_sender) >= 3 and (excel_sender in skey or skey in excel_sender)):
                        locked_cand = c if c.get("is_locked") else None
                        if locked_cand:
                            matched = locked_cand
                            status = "duplicate"
                            method = "نام جزئی مشابه + مبلغ یکسان (تکراری)"
                            matched["duplicate_lock_text"] = matched.get("lock_text", "قبلاً تطبیق شده")
                        else:
                            matched = c
                            method, status = f"نام جزئی مشابه + مبلغ یکسان", "review"
                        break

        # ── 3. Removed: Smart amount-only match ──
        # ── 5. Setup bank_rows array for multiple matches ──
        bank_rows_list = []
        if status == "multiple" and matched and "all_matched_rows" in matched:
             bank_rows_list = matched["all_matched_rows"]
        elif matched and matched.get("row_num"):
             bank_rows_list = [matched.get("row_num")]

        results.append({
            "idx":          len(results) + 1,
            "date":         r.get("date", ""),
            "doc_num":      r.get("doc_num", ""),
            "doc_type":     r.get("doc_type", ""),
            "amount":       int(amount) if amount else 0,
            "codes":        ", ".join(codes),
            "sender":       sender,
            "customer_name": r.get("customer_name", ""),
            "desc":         r.get("desc", ""),
            "found":        matched is not None,
            "status":       status,
            "match_method": method,
            "bank_ref":     matched.get("ref", "") if matched else "",
            "bank_date":    matched.get("date", "") if matched else "",
            "bank_row":     matched.get("row_num", "") if matched else "",
            "bank_rows":    bank_rows_list,
            "bank_sender":  matched.get("sender", "") if matched else "",
            "duplicate_lock_text": matched.get("duplicate_lock_text", "") if matched else "",
        })

    # Cross-file duplicate detection: downgrade status for rows matched in previous HTMLs
    if already_used_rows:
        for r in results:
            row_num = r.get("bank_row")
            if row_num and row_num in already_used_rows and r["status"] in ("exact", "review"):
                r["status"] = "duplicate"
                r["match_method"] = r.get("match_method", "") + " — در فایل HTML قبلی تطبیق شده"

    return results

# ──────────────────────────────────────────────────────────────────────────────
# FastAPI Routes
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    if not request.session.get("authenticated"):
        return RedirectResponse(url="/login", status_code=303)
    html_path = BASE_DIR / "index.html"
    return html_path.read_text(encoding="utf-8")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if request.session.get("authenticated"):
        return RedirectResponse(url="/")
    html_path = BASE_DIR / "login.html"
    return html_path.read_text(encoding="utf-8")

@app.post("/api/login")
async def do_login(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == ADMIN_USER and password == ADMIN_PASS:
        request.session["authenticated"] = True
        return JSONResponse(content={"ok": True})
    return JSONResponse(status_code=401, content={"ok": False, "error": "Invalid credentials"})

@app.post("/api/logout")
async def do_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

# ──────────────────────────────────────────────────────────────────────────────
# Session Excel Upload & Status
# ──────────────────────────────────────────────────────────────────────────────

@app.post("/api/upload-excel")
async def upload_excel_to_session(
    request: Request,
    excel_file: UploadFile = File(...),
):
    check_auth(request)
    _ensure_session_id(request)
    sid = _session_key(request)
    excel_bytes = await excel_file.read()
    try:
        bank_txns = parse_excel(excel_bytes, excel_file.filename or "bank.xls")
    except Exception as e:
        raise HTTPException(500, f"خطا در پردازش Excel: {e}")
    EXCEL_SESSIONS[sid] = {
        "bank_txns":       bank_txns,
        "excel_bytes":     excel_bytes,
        "excel_filename":  excel_file.filename or "bank.xls",
        "matched_row_nums": set(),
        "html_files":      [],
    }
    return JSONResponse({"ok": True, "filename": excel_file.filename, "rows": len(bank_txns)})


@app.get("/api/session-status")
async def session_status(request: Request):
    check_auth(request)
    sid = _session_key(request)
    sess = EXCEL_SESSIONS.get(sid)
    if not sess:
        return JSONResponse({"ok": True, "has_excel": False})
    return JSONResponse({
        "ok":           True,
        "has_excel":    True,
        "filename":     sess["excel_filename"],
        "bank_rows":    len(sess["bank_txns"]),
        "matched_count": len(sess["matched_row_nums"]),
        "html_files":   sess["html_files"],
    })


@app.delete("/api/clear-session")
async def clear_session(request: Request):
    check_auth(request)
    sid = _session_key(request)
    if sid in EXCEL_SESSIONS:
        del EXCEL_SESSIONS[sid]
    return JSONResponse({"ok": True})

class LegacySSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = create_urllib3_context()
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        # This allows older insecure ciphers that the local software might be using
        context.set_ciphers('DEFAULT@SECLEVEL=0')
        kwargs['ssl_context'] = context
        return super(LegacySSLAdapter, self).init_poolmanager(*args, **kwargs)

@app.post("/api/test-connection")
async def test_api_connection(request: Request, data: dict = Body(...)):
    check_auth(request)
    url = data.get("url")
    method = data.get("method", "GET").upper()
    headers = data.get("headers", {})
    payload = data.get("payload")
    
    if not url:
        raise HTTPException(status_code=400, detail="آدرس API وارد نشده است.")
        
    try:
        # Suppress insecure request warnings
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        session = requests.Session()
        session.mount('https://', LegacySSLAdapter())
        
        if method == "GET":
            # payload is used as query params
            response = session.get(url, headers=headers, params=payload, timeout=15, verify=False)
        elif method == "POST":
            # payload is used as json body
            response = session.post(url, headers=headers, json=payload, timeout=15, verify=False)
        else:
            raise HTTPException(status_code=400, detail="متد نامعتبر است.")
            
        try:
            resp_json = response.json()
        except Exception:
            resp_json = {"raw_text": response.text[:2000] + ("..." if len(response.text)>2000 else "")}
            
        return {
            "ok": True,
            "status_code": response.status_code,
            "response": resp_json
        }
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": str(e)}

@app.post("/analyze-from-api")
async def analyze_from_api(
    request: Request,
    excel_file:      UploadFile = File(...),
    loaded_receipts: str        = Form(...),
    selected_banks:  str        = Form(""),
    credit_only:     str        = Form("true"),
    use_tracking:    str        = Form("true"),
    use_name:        str        = Form("true"),
    use_amount:      str        = Form("true"),
    tx_type_filter:  str        = Form("all"),
    use_date:        str        = Form("false"),
):
    check_auth(request)
    excel_bytes = await excel_file.read()

    allowed_banks = [b.strip() for b in selected_banks.split(",") if b.strip()]
    is_all_banks = "ALL" in allowed_banks

    try:
        asnad_raw = json.loads(loaded_receipts)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid loaded receipts JSON data")

    if isinstance(asnad_raw, dict) and "ERROR" in asnad_raw:
        raise HTTPException(status_code=400, detail=asnad_raw.get("ERROR", "خطا از ته حساب"))

    # ── Normalize DoListAsnad → pdf_rows ─────────────────────────────────────────
    # Each record has: Tarikh, NO (نوع سند), Mali (مالی), Sh_Factor, Sharh1, MCode, Bank_Name
    pdf_rows = []
    
    # asnad_raw could be a dict { "1": {...} } or a list [{...}]
    raw_list = []
    if isinstance(asnad_raw, dict):
        raw_list = list(asnad_raw.values())
    elif isinstance(asnad_raw, list):
        raw_list = asnad_raw

    for rec in raw_list:
        if not isinstance(rec, dict):
            continue
        
        # Note: bank and doc type filtering is already done on the frontend before sending
        no_sanad = str(rec.get("NO", "")).strip()
        mali = rec.get("Mali")
        tarikh = str(rec.get("Tarikh", "")).strip()
        sh = str(rec.get("Sh_Factor", "")).strip()
        sharh1 = str(rec.get("Sharh1") or "").strip()
        sharh2 = str(rec.get("Sharh2") or "").strip()
        sharh = sharh1 or sharh2
        logger.info(f"API Record -> No: {no_sanad}, Mali: {mali}, Sh: {sh}, Sharh: {sharh}")
        mcode = str(rec.get("MCode", "")).strip()

        # Classify: if Mali > 0 = بستانکار, Mali < 0 = بدهکار
        try:
            mali_str = str(mali).replace(",", "").strip()
            mali_val = float(mali_str)
        except (TypeError, ValueError):
            mali_val = 0

        doc_type = "بستانکار" if mali_val > 0 else "بدهکار"
        amount = abs(mali_val)

        # Extract 4-digit tracking codes from Sharh AND Sh_Factor
        # Sh_Factor often contains the receipt sequence number on the bank statement
        import re as _re
        code_src = f"{sharh} {sharh2} {sh}"
        codes = list(set(_re.findall(r'\b\d{4}\b', code_src)))
        # Also add Sh_Factor itself as a code if it's purely numeric (4-5 digits)
        if sh and sh.isdigit() and 4 <= len(sh) <= 5:
            codes.append(sh)
        sender = sharh if sharh else (sh if sh else mcode)

        pdf_rows.append({
            "date":     tarikh,
            "doc_num":  sh,
            "doc_type": doc_type,
            "amount":   amount,
            "codes":    codes,
            "sender":   sender,
            "desc":     f"{sharh} {sharh2}".strip(),
            "no":       no_sanad,
        })

    # ── Parse Excel ──────────────────────────────────────────────────────────────
    try:
        bank_txns = parse_excel(excel_bytes, excel_file.filename or "bank.xls")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"خطا در خواندن فایل Excel: {str(e)}")

    # ── Match ────────────────────────────────────────────────────────────────────
    results = match_receipts(
        pdf_rows,
        bank_txns,
        credit_only=credit_only.lower() == "true",
        use_tracking=use_tracking.lower() == "true",
        use_name=use_name.lower() == "true",
        use_amount=use_amount.lower() == "true",
        tx_type_filter=tx_type_filter,
        use_date=use_date.lower() == "true",
    )

    found     = sum(1 for r in results if r["status"] == "exact")
    review    = sum(1 for r in results if r["status"] == "review")
    multiple  = sum(1 for r in results if r["status"] == "multiple")
    duplicate = sum(1 for r in results if r["status"] == "duplicate")
    not_found = sum(1 for r in results if r["status"] == "notfound")

    return {
        "ok": True,
        "pdf_total":  len(pdf_rows),
        "bank_total": len(bank_txns),
        "found":      found,
        "review":     review,
        "multiple":   multiple,
        "duplicate":  duplicate,
        "not_found":  not_found,
        "results":    results,
    }

@app.post("/analyze")
async def analyze(
    request:     Request,
    pdf_file:    UploadFile = File(...),
    credit_only: str        = Form("true"),
    use_tracking:str        = Form("true"),
    use_name:    str        = Form("true"),
    use_amount:  str        = Form("true"),
    tx_type_filter: str     = Form("all"),
    use_date:    str        = Form("false"),
):
    check_auth(request)
    sid = _session_key(request)
    sess = EXCEL_SESSIONS.get(sid)
    if not sess:
        raise HTTPException(400, "فایل اکسل بارگذاری نشده. ابتدا فایل حساب بانکی (اکسل) را آپلود کنید.")

    bank_txns    = sess["bank_txns"]
    excel_bytes  = sess["excel_bytes"]
    excel_filename = sess["excel_filename"]
    already_used = sess["matched_row_nums"]

    pdf_bytes   = await pdf_file.read()

    filename = (pdf_file.filename or "").lower()
    is_html = filename.endswith(".html") or filename.endswith(".htm")

    try:
        if is_html:
            pdf_rows = parse_html(pdf_bytes)
        else:
            pdf_rows = parse_pdf(pdf_bytes)
    except Exception as e:
        logger.exception(f"Exception parsing {filename}")
        raise HTTPException(500, f"خطا در پردازش فایل: {e}")

    results = match_receipts(
        pdf_rows, bank_txns,
        credit_only    = credit_only.lower() == "true",
        use_tracking   = use_tracking.lower() == "true",
        use_name       = use_name.lower()     == "true",
        use_amount     = use_amount.lower()   == "true",
        tx_type_filter = tx_type_filter.lower(),
        use_date       = use_date.lower()     == "true",
        already_used_rows = already_used,
    )

    # Update session: add newly matched row numbers + record this HTML file
    new_matched = set()
    for r in results:
        if r["status"] in ("exact", "review"):
            if r.get("bank_row"):
                new_matched.add(r["bank_row"])
        elif r["status"] == "multiple":
            for br in (r.get("bank_rows") or []):
                new_matched.add(br)
    sess["matched_row_nums"].update(new_matched)
    sess["html_files"].append({
        "name":    pdf_file.filename or "ناشناخته",
        "matched": len(new_matched),
        "total":   len(pdf_rows),
    })

    # ── Excel Locking Generation ──
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import PatternFill
        import io
        import os

        # Get 1-based row numbers from matched results
        new_matched_rows = set()
        multiple_matched_rows = set()
        
        for r in results:
            if r["status"] in ("exact", "review", "duplicate"):
                if r.get("bank_row"):
                    new_matched_rows.add(r["bank_row"])
            elif r["status"] == "multiple":
                if r.get("bank_rows"):
                    for b_row in r["bank_rows"]:
                        multiple_matched_rows.add(b_row)

        all_matched_rows = set(new_matched_rows)
        # Add multiple matched rows to the all set so we can color them
        all_matched_rows.update(multiple_matched_rows)
        
        # Keep track of old lock texts to preserve them
        old_locks = {}
        
        # Also include previously locked rows from the bank txns so they don't lose their color
        # because pandas rebuilds the file without preserving original formatting.
        for t in bank_txns:
            if t.get("is_locked") and t.get("row_num"):
                # +1 because pandas to_excel without header/index makes 0-indexed rows 1-indexed in openpyxl,
                # but tx['row_num'] is already aligned to openpyxl 1-based indexing in parse_excel
                all_matched_rows.add(t["row_num"])
                if t.get("lock_text"):
                    old_locks[t["row_num"]] = t["lock_text"]

        if all_matched_rows:
            wb = None
            try:
                # Attempt to load directly with openpyxl (best for preserving formatting)
                wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            except Exception as e:
                logger.warning(f"Native openpyxl load failed ({e}), falling back to pandas conversion...")
                # Fallback: Read raw bytes with Pandas and convert to standard XLSX in memory
                # (Often needed for older .xls bounds or HTML tables disguised as excel)
                df = pd.read_excel(io.BytesIO(excel_bytes), header=None)
                xlsx_buffer = io.BytesIO()
                with pd.ExcelWriter(xlsx_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, header=False)
                xlsx_buffer.seek(0)
                wb = openpyxl.load_workbook(xlsx_buffer)

            ws = wb.active
            ws.sheet_view.rightToLeft = True
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

            # Pandas drops the 1-based indexing, so row N in pandas is row N in openpyxl
            
            pdf_base_name = Path(pdf_file.filename or "ناشناس").stem
            new_lock_msg = f"تطبیق شده - {pdf_base_name}"
            
            # Define violet/purple color for multiple (ambiguous) matches
            pink_fill = PatternFill(start_color="FFDDA0DD", end_color="FFDDA0DD", fill_type="solid")
            
            for r_idx in all_matched_rows:
                # If it's a multiple match, color it pink, otherwise yellow
                current_fill = pink_fill if r_idx in multiple_matched_rows else yellow_fill
                
                # Check for existing lock to append "و"
                existing = old_locks.get(r_idx, "")
                if existing and "تطبیق شده" in existing and pdf_base_name not in existing:
                    # Append the new file if it's already locked by another file
                    msg = f"{existing} و {pdf_base_name}"
                else:
                    msg = existing if existing else new_lock_msg
                
                ws.cell(row=r_idx, column=15, value=msg) 
                
                # Color the row
                for col in range(1, 16):
                    ws.cell(row=r_idx, column=col).fill = current_fill

            # Save to memory instead of writing to a hardcoded Windows path
            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            sess["colored_bytes"] = final_buffer.getvalue()
            
            logger.info(f"Generated locked Excel with {len(all_matched_rows)} rows highlighted and saved to session memory.")
    except Exception as e:
        logger.error(f"Failed to generate locked excel: {e}")

    found_count = sum(1 for r in results if r["status"] == "exact")
    review_count= sum(1 for r in results if r["status"] == "review")
    dupl_count  = sum(1 for r in results if r["status"] == "duplicate")
    mult_count  = sum(1 for r in results if r["status"] == "multiple")
    
    # We map 'found' technically as exact matches, 'review' as review
    # The frontend uses status to categorize the UI boxes
    return JSONResponse({
        "ok":         True,
        "total":      len(results),
        "found":      found_count,
        "review":     review_count,
        "duplicate":  dupl_count,
        "multiple":   mult_count,
        "not_found":  len(results) - found_count - review_count - dupl_count - mult_count,
        "bank_total": len(bank_txns),
        "pdf_total":  len(pdf_rows),
        "results":    results,
        "debug": {
            "pdf_rows_sample":  pdf_rows[:5],
            "bank_txns_sample": bank_txns[:5],
        }
    })

@app.get("/api/download-excel")
async def download_session_excel(request: Request):
    """Download the cumulative edited/colored bank Excel for the current session."""
    check_auth(request)
    session_id = _session_key(request)
    if not session_id or session_id not in EXCEL_SESSIONS:
        return JSONResponse({"ok": False, "detail": "اکسل در حافظه یافت نشد"})
        
    session_data = EXCEL_SESSIONS[session_id]
    colored_bytes = session_data.get("colored_bytes")
    # If not completely analyzed, or if it's there but blank
    if not colored_bytes:
        return JSONResponse({"ok": False, "detail": "اکسل هنوز رنگ‌آمیزی نشده (اول باید حداقل یک بار تطبیق انجام بشه)"})
        
    from pathlib import Path
    import urllib.parse
    
    orig_name = session_data.get("filename", "bank_statement.xlsx")
    base_name = Path(orig_name).stem
    new_filename = f"{base_name} - فیش‌های بررسی شده.xlsx"
    encoded_name = urllib.parse.quote(new_filename)
    
    return StreamingResponse(
        io.BytesIO(colored_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"
        }
    )

@app.get("/health")
async def health():
    return {"ok": True}

if __name__ == "__main__":
    import sys, io
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    print("\n" + "="*60)
    print("  Receipt Checker - http://localhost:8765")
    print("="*60 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=8765, log_level="info")
