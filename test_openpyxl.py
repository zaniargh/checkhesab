import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

df = pd.read_excel('d:/Checkhesab/ایران 1-6.xls', header=None)
out_path = 'd:/Checkhesab/output_test.xlsx'
df.to_excel(out_path, index=False, header=False)

# Now open with openpyxl to color
wb = openpyxl.load_workbook(out_path)
ws = wb.active
yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

for col in range(1, 15):
    ws.cell(row=10, column=col).fill = yellow
    
wb.save(out_path)
print("Saved to", out_path)
